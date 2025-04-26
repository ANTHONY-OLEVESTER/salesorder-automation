import os
import pandas as pd
import requests
import json
from dotenv import load_dotenv
from openpyxl import Workbook

# Load environment variables
load_dotenv()

EXCEL_FILE = "salesorders_Today.xlsx"
SALESORDERS_URL = "https://www.**Redacted**apis.in/inventory/v1/salesorders"
SALESORDER_DETAIL_URL = "https://www.**Redacted**apis.in/inventory/v1/salesorders/{}"
Z_ORG_ID = os.getenv("**Redacted**_ORG_ID")
Z_AUTH_TOKEN = os.getenv("**Redacted**_ACCESS_TOKEN")

# Load mapping
with open("map.json", "r", encoding="utf-8") as f:
    FIELD_MAP = json.load(f)

REVERSE_MAP = {v: k for k, v in FIELD_MAP.items()}
HEADERS = list(FIELD_MAP.values())

headers = {
    "Authorization": f"**Redacted**-oauthtoken {Z_AUTH_TOKEN}",
    "X-com-**Redacted**-inventory-organizationid": Z_ORG_ID
}

# 📦 Step 1: Fetch all sales orders
print("📦 Fetching sales orders...")
sales_response = requests.get(SALESORDERS_URL, headers=headers)

if sales_response.status_code != 200:
    print("❌ Failed to fetch sales orders", sales_response.text)
    exit()

salesorders = sales_response.json().get("salesorders", [])
print(f"✅ Retrieved {len(salesorders)} sales orders")

# 🧩 Step 2: Fetch full details for each sales order and expand line items
expanded_rows = []

for order in salesorders:
    salesorder_id = order.get("salesorder_id")
    detail_url = SALESORDER_DETAIL_URL.format(salesorder_id)
    detail_response = requests.get(detail_url, headers=headers)

    if detail_response.status_code != 200:
        print(f"⚠️ Failed to fetch details for order {salesorder_id}")
        continue

    order_detail = detail_response.json().get("salesorder", {})

    # Flatten custom fields
    custom_fields = order_detail.pop("custom_fields", [])
    for field in custom_fields:
        order_detail[f"custom_fields.{field.get('label')}"] = field.get("value")

    # Flatten billing/shipping address
    for addr_key in ["billing_address", "shipping_address"]:
        addr = order_detail.get(addr_key, {})
        for k, v in addr.items():
            order_detail[f"{addr_key}.{k}"] = v

    # Handle line items by duplicating sales order per item
    line_items = order_detail.pop("line_items", [])
    if not line_items:
        expanded_rows.append(order_detail)
    else:
        for item in line_items:
            row = order_detail.copy()
            for k, v in item.items():
                if isinstance(v, dict):
                    for subk, subv in v.items():
                        row[f"item.{k}.{subk}"] = subv
                else:
                    row[f"item.{k}"] = v
            expanded_rows.append(row)

# 🗂 Step 3: Map and export
wb = Workbook()
del wb[wb.sheetnames[0]]  # remove default sheet

ws = wb.create_sheet(title="SalesOrders")
ws.append(HEADERS)

for row in expanded_rows:
    mapped = []
    for header in HEADERS:
        api_field = REVERSE_MAP.get(header)
        value = row.get(api_field, "")
        mapped.append(value)
    ws.append(mapped)

wb.save(EXCEL_FILE)
print(f"✅ Sales order data mapped and saved to {EXCEL_FILE}")

import os
import re
import mysql.connector
import pandas as pd
import json
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "database": os.getenv("DB_NAME"),
    "port": int(os.getenv("DB_PORT", 3306))
}

EXCEL_FILE = "salesorders_Today.xlsx"
TABLE_NAME = "sales_orders"

# Load mapping from map_sql.json
with open("map_sql.json", "r", encoding="utf-8") as f:
    FIELD_MAP = json.load(f)



def load_excel(filename):
    print(f"📥 Reading {filename}...")
    df = pd.read_excel(filename)
    print(f"✅ Loaded {len(df)} rows with {len(df.columns)} columns.")
    return df


def connect_to_db():
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        print("✅ Connected to database")
        return conn
    except mysql.connector.Error as err:
        print("❌ DB Connection Error:", err)
        return None


def upload_to_db(df, conn):
    cursor = conn.cursor()

    # Get actual column names from the SQL table
    cursor.execute(f"SHOW COLUMNS FROM {TABLE_NAME}")
    db_cols = set(row[0].lower() for row in cursor.fetchall())

    for index, row in df.iterrows():
        valid_cols = []
        values = []
        seen = set()

        for db_col, excel_col in FIELD_MAP.items():
            if db_col.lower() in db_cols and db_col.lower() not in seen:
                valid_cols.append(db_col)
                value = row.get(excel_col)
                values.append(None if pd.isna(value) else str(value))
                seen.add(db_col.lower())


        if not valid_cols:
            print(f"⚠️ Row {index} skipped: No matching columns found in DB for this row")
            continue


        placeholders = ["%s"] * len(valid_cols)
        sql = f"""
        INSERT INTO {TABLE_NAME} ({', '.join(valid_cols)})
        VALUES ({', '.join(placeholders)})
        ON DUPLICATE KEY UPDATE {', '.join([f'{col}=VALUES({col})' for col in valid_cols])}
        """
        try:
            cursor.execute(sql, values)
        except Exception as e:
            print(f"❌ Failed to insert row {index}: {e}")

    conn.commit()
    print(f"✅ Uploaded {len(df)} rows to `{TABLE_NAME}`")


if __name__ == "__main__":
    df = load_excel(EXCEL_FILE)
    conn = connect_to_db()
    if conn:
        upload_to_db(df, conn)
        conn.close()
